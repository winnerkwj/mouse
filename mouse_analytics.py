"""마우스 사용 분석기 (히트맵 + 이동량 + 클릭수 + 엑셀 자동저장).

회사에서 어떤 SW를 사용하든, 그 동안의 마우스 활동을 전역으로 측정해
초보자도 쉽게 분석 결과를 얻도록 만든 단일 창 도구다. 측정 항목:

  · 클릭 위치 히트맵 (빈 캔버스 / 화면 캡처 배경 중 선택)
  · 마우스 이동 거리 (픽셀 + 물리 cm)
  · 버튼별 클릭 횟수 + 스크롤
  · 위 기록을 엑셀(.xlsx)로 주기적·자동 저장 (summary / events / heatmap 시트)

전역 단축키:
    Ctrl+Shift+F9  : 녹화 시작/정지
    Ctrl+Shift+F10 : 히트맵 생성 + 엑셀 저장

기존 도구(mouse_click_move2.py / click_update_v1.0.9.py / mouse_distance1.0.6.py)의
검증된 패턴을 재사용해 하나로 통합했다. Windows 전용.

자기 UI 클릭 제외: 우리 창/다이얼로그를 누른 클릭은 분석 데이터에서 제외해
잘못된 기록이 남지 않게 한다(_own_rect / _suppress). 녹화 중에는 단축키로
제어하거나 '시작 시 창 최소화'를 쓰면 우리 UI 클릭 자체가 발생하지 않는다.
"""

import os
import math
import time
import shutil
import logging
import datetime
import tempfile
import threading
import ctypes
from ctypes import wintypes
from types import SimpleNamespace

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import numpy as np
from PIL import Image, ImageGrab
from pynput import mouse, keyboard
from pynput.mouse import Button
from screeninfo import get_monitors

import openpyxl
from openpyxl.drawing.image import Image as XLImage

# --- 상수 -------------------------------------------------------------------
DEFAULT_DPI = 96             # 화면 크기 미입력 시 기본 DPI
INCH_TO_MM = 25.4
MM_TO_CM = 0.1
SCROLL_STEP_MM = 15          # 스크롤 한 칸의 추정 이동 거리(mm)
POINT_CAP = 5000             # 히트맵 포인트/이벤트 로그 최대 보관 수(메모리 한계)
MOVE_MIN_INTERVAL_S = 0.10   # 이동 이벤트 코얼레싱: 최소 시간 간격
MOVE_MIN_DIST_PX = 50        # 이동 이벤트 코얼레싱: 최소 이동 거리
DEFAULT_AUTOSAVE_S = 30
GAUSS_SIGMA = 30             # 히트맵 가우시안 블러 반경
HEAT_ALPHA_GAMMA = 0.55      # 히트맵 알파 감마(작을수록 중간 밀도도 잘 보임)
HEAT_MAX_ALPHA = 0.85        # 핫스팟 최대 불투명도
BLUE_BASE_DEFAULT = 0.13     # 배경에 깔리는 옅은 파란 기운(0=없음) — UI 슬라이더로 조절
BLUE_BASE_RGB = (70, 110, 225)
APP_DIR_NAME = "MouseAnalytics"


def _gaussian_blur(arr, sigma):
    """numpy 만으로 float 분리형 가우시안 블러.

    Pillow 의 uint8 GaussianBlur 은 sigma 가 크면 희소한 클릭의 블러 값이 1 미만으로
    뭉개져 전부 0 이 되어 히트맵이 텅 비는 버그가 있었다. float 로 블러해 작은 값도
    보존한다(scipy.ndimage.gaussian_filter 과 동일한 효과, 의존성 없이).
    """
    arr = np.asarray(arr, dtype=np.float32)
    radius = int(max(1, round(3.0 * sigma)))
    xs = np.arange(-radius, radius + 1, dtype=np.float32)
    k = np.exp(-(xs * xs) / (2.0 * sigma * sigma)).astype(np.float32)
    k /= k.sum()
    out = np.empty_like(arr)
    for i in range(arr.shape[0]):          # 가로 방향
        out[i] = np.convolve(arr[i], k, mode="same")
    for j in range(out.shape[1]):          # 세로 방향
        out[:, j] = np.convolve(out[:, j], k, mode="same")
    return out


def _turbo_rgb(t):
    """[0,1] 2D 배열 -> (H,W,3) float RGB. Google 'Turbo' 컬러맵 다항식 근사.

    jet 보다 색 띠(rings)가 적고 매끄러워 위치 표현이 자연스럽다. matplotlib 없이
    numpy 만으로 계산한다(PyInstaller 패키징 경량/안정).
    """
    t = np.clip(t, 0.0, 1.0)
    t2, t3, t4, t5 = t * t, t ** 3, t ** 4, t ** 5
    r = 0.13572138 + 4.61539260 * t - 42.66032258 * t2 + 132.13108234 * t3 - 152.94239396 * t4 + 59.28637943 * t5
    g = 0.09140261 + 2.19418839 * t + 4.84296658 * t2 - 14.18503333 * t3 + 4.27729857 * t4 + 2.82956604 * t5
    b = 0.10667330 + 12.64194608 * t - 60.58204836 * t2 + 110.36276771 * t3 - 89.90310912 * t4 + 27.34824973 * t5
    return np.clip(np.stack([r, g, b], axis=-1), 0.0, 1.0)


class MouseAnalytics(tk.Tk):
    def __init__(self):
        super().__init__()

        # 임시 폴더 + 로깅 (mouse_click_move2.py:32-45)
        self.temp_dir = os.path.join(tempfile.gettempdir(), "mouse_analytics_app")
        try:
            os.makedirs(self.temp_dir, exist_ok=True)
        except OSError:
            self.temp_dir = tempfile.gettempdir()
        logging.basicConfig(
            filename=os.path.join(self.temp_dir, "error.log"),
            level=logging.INFO,
            format="%(asctime)s %(levelname)s: %(message)s",
        )
        logging.info("Mouse analytics started.")

        # 모니터 열거 (mouse_click_move2.py:47)
        try:
            self.monitors = list(get_monitors())
        except Exception as e:
            logging.error("get_monitors failed: %s", e)
            self.monitors = []
        if not self.monitors:
            self.monitors = [SimpleNamespace(x=0, y=0, width=1920, height=1080, name="기본")]

        # 공유 상태 (self.lock으로 보호)
        self.lock = threading.Lock()
        self.click_positions = []   # 히트맵용 (모니터 상대좌표), POINT_CAP 캡
        self.click_counts = {Button.left: 0, Button.right: 0, Button.middle: 0}
        self.total_distance_mm = 0.0
        self.scroll_count = 0       # 스크롤 칸 수(누적)
        self.events = []            # (ts, type, button, x, y, monitor_idx, dist_px)

        # 세션/리스너 상태
        self.is_recording = False
        self.listener = None
        self.session_start = None
        self.session_file = None
        self._last_heatmap_png = None
        self._session_shot = None        # 현재 단계의 배경(스크린샷). 단계가 바뀔 때마다 캡처
        self._steps = []                 # 완료된 단계 기록(각 단계 = 히트맵 PNG + 통계)
        self._step_no = 1                # 현재 단계 번호
        self._step_start = None          # 현재 단계 시작 시각

        # 녹화 시작 시 고정되는 캐시값(리스너 스레드가 읽음)
        self._active_monitor = None
        self._active_monitor_idx = 0
        self._ppm = DEFAULT_DPI / INCH_TO_MM   # pixels per mm
        self._rec_move = True
        self._rec_scroll = True

        # 이동 이벤트 코얼레싱 상태
        self._last_move_pos = None
        self._last_sample_pos = None
        self._last_sample_t = 0.0
        self._win_dist = 0.0

        # 자기 UI 클릭 제외용
        self._own_rect = None       # (x1,y1,x2,y2) 또는 None
        self._suppress = False      # 모달 다이얼로그 표시 중 True

        self._autosave_remaining = DEFAULT_AUTOSAVE_S

        self.build_ui()
        # 창 이동/리사이즈 시 자기 영역 갱신(메인 스레드)
        self.bind("<Configure>", lambda e: self._update_own_rect())

        # 전역 단축키 (mouse_click_move2.py:71-86 재사용). 콜백은 메인 스레드로 마샬링.
        try:
            self.keyboard_listener = keyboard.GlobalHotKeys({
                "<ctrl>+<shift>+<f9>": lambda: self.after(0, self.toggle_recording),
                "<ctrl>+<shift>+<f10>": lambda: self.after(0, self.advance_step),
            })
            self.keyboard_listener.start()
        except Exception as e:
            logging.error("Keyboard listener start failed: %s", e)
            self.keyboard_listener = None

        self.after(1000, self.tick)

        messagebox.showinfo(
            "안내",
            "이 프로그램은 마우스 클릭/이동/스크롤을 전역으로 측정하고\n"
            "결과를 엑셀과 히트맵으로 자동 저장합니다.\n\n"
            "· 백신/권한 문제 시 예외 등록 또는 관리자 권한 실행을 권장합니다.\n"
            "· 녹화 중에는 단축키(Ctrl+Shift+F9)로 제어하면 됩니다.",
        )

    # --- UI -----------------------------------------------------------------
    def build_ui(self):
        self.title("마우스 사용 분석기")
        self.resizable(False, False)
        pad = {"padx": 8, "pady": 2}

        self.status_label = tk.Label(self, text="○ 대기 중", fg="gray",
                                     font=("맑은 고딕", 11, "bold"))
        self.status_label.pack(pady=(10, 4))

        self.start_button = tk.Button(
            self, text="▶ 녹화 시작 (Ctrl+Shift+F9)",
            font=("맑은 고딕", 12, "bold"), bg="#5cb85c", fg="white",
            width=28, height=2, command=self.toggle_recording,
        )
        self.start_button.pack(pady=6)

        stat = tk.LabelFrame(self, text="실시간 측정", font=("맑은 고딕", 9))
        stat.pack(fill="x", padx=10, pady=6)
        self.clicks_label = tk.Label(stat, text="클릭  총 0  (좌 0 · 우 0 · 휠 0)",
                                     anchor="w", font=("맑은 고딕", 10))
        self.clicks_label.pack(fill="x", **pad)
        self.distance_label = tk.Label(stat, text="이동 거리  0 px  /  0.0 cm",
                                       anchor="w", font=("맑은 고딕", 10))
        self.distance_label.pack(fill="x", **pad)
        self.scroll_label = tk.Label(stat, text="스크롤  0 칸",
                                     anchor="w", font=("맑은 고딕", 10))
        self.scroll_label.pack(fill="x", **pad)
        self.time_label = tk.Label(stat, text="경과 시간  00:00:00",
                                   anchor="w", font=("맑은 고딕", 10))
        self.time_label.pack(fill="x", **pad)
        self.monitor_label = tk.Label(stat, text="모니터  —",
                                      anchor="w", font=("맑은 고딕", 9), fg="gray")
        self.monitor_label.pack(fill="x", **pad)

        tk.Label(self,
                 text="단축키:  시작/정지 Ctrl+Shift+F9   ·   다음 단계 저장 Ctrl+Shift+F10",
                 font=("맑은 고딕", 8), fg="#555").pack(pady=(0, 4))

        cfg = tk.LabelFrame(self, text="설정", font=("맑은 고딕", 9))
        cfg.pack(fill="x", padx=10, pady=6)

        tk.Label(cfg, text="모니터").grid(row=0, column=0, sticky="w", **pad)
        self.monitor_combo = ttk.Combobox(cfg, values=self._monitor_names(),
                                          state="readonly", width=30)
        self.monitor_combo.current(self._default_monitor_index())
        self.monitor_combo.grid(row=0, column=1, columnspan=3, sticky="w", **pad)

        tk.Label(cfg, text="해상도(px)").grid(row=1, column=0, sticky="w", **pad)
        self.res_w_entry = tk.Entry(cfg, width=7)
        self.res_w_entry.grid(row=1, column=1, sticky="w", **pad)
        tk.Label(cfg, text="×").grid(row=1, column=2, sticky="w")
        self.res_h_entry = tk.Entry(cfg, width=7)
        self.res_h_entry.grid(row=1, column=3, sticky="w", **pad)

        tk.Label(cfg, text="대각선(inch)").grid(row=2, column=0, sticky="w", **pad)
        self.diag_entry = tk.Entry(cfg, width=7)
        self.diag_entry.grid(row=2, column=1, sticky="w", **pad)
        tk.Label(cfg, text="← 정확한 cm 환산용", fg="gray",
                 font=("맑은 고딕", 8)).grid(row=2, column=2, columnspan=2, sticky="w")

        tk.Label(cfg, text="자동저장(초)").grid(row=3, column=0, sticky="w", **pad)
        self.autosave_entry = tk.Entry(cfg, width=7)
        self.autosave_entry.grid(row=3, column=1, sticky="w", **pad)

        tk.Label(cfg, text="저장 폴더").grid(row=4, column=0, sticky="w", **pad)
        self.folder_entry = tk.Entry(cfg, width=28)
        self.folder_entry.grid(row=4, column=1, columnspan=2, sticky="we", **pad)
        tk.Button(cfg, text="찾아보기", command=self._browse_folder).grid(
            row=4, column=3, sticky="w", **pad)

        self.move_var = tk.BooleanVar(value=True)
        self.scroll_var = tk.BooleanVar(value=True)
        self.autosave_var = tk.BooleanVar(value=True)
        self.minimize_var = tk.BooleanVar(value=True)
        tk.Checkbutton(cfg, text="이동 기록", variable=self.move_var).grid(
            row=5, column=0, sticky="w", **pad)
        tk.Checkbutton(cfg, text="스크롤 기록", variable=self.scroll_var).grid(
            row=5, column=1, sticky="w", **pad)
        tk.Checkbutton(cfg, text="자동저장", variable=self.autosave_var).grid(
            row=5, column=2, sticky="w", **pad)
        tk.Checkbutton(cfg, text="시작 시 창 최소화", variable=self.minimize_var).grid(
            row=6, column=0, columnspan=2, sticky="w", **pad)
        self.bg_include_var = tk.BooleanVar(value=True)
        tk.Checkbutton(cfg, text="화면 배경 포함", variable=self.bg_include_var).grid(
            row=6, column=2, columnspan=2, sticky="w", **pad)

        tk.Label(cfg, text="파란 배경 세기").grid(row=7, column=0, sticky="w", **pad)
        self.blue_base_var = tk.DoubleVar(value=BLUE_BASE_DEFAULT)
        tk.Scale(cfg, variable=self.blue_base_var, from_=0.0, to=0.4, resolution=0.01,
                 orient="horizontal", showvalue=True, length=150).grid(
            row=7, column=1, columnspan=3, sticky="w", **pad)

        act = tk.Frame(self)
        act.pack(fill="x", padx=10, pady=(4, 10))
        tk.Button(act, text="다음 단계 저장 (Ctrl+Shift+F10)",
                  command=self.advance_step).pack(side="left", padx=4)
        tk.Button(act, text="폴더 열기", command=self._open_folder).pack(side="left", padx=4)

        # 기본값 채우기 (주 모니터 기준)
        m0 = self.monitors[self._default_monitor_index()]
        self.res_w_entry.insert(0, str(getattr(m0, "width", 1920)))
        self.res_h_entry.insert(0, str(getattr(m0, "height", 1080)))
        self.diag_entry.insert(0, "24")
        self.autosave_entry.insert(0, str(DEFAULT_AUTOSAVE_S))
        self.folder_entry.insert(
            0, os.path.join(os.path.expanduser("~"), "Desktop", APP_DIR_NAME))

    def _monitor_names(self):
        return [f"모니터 {i}: ({m.width}x{m.height}) at ({m.x},{m.y})"
                for i, m in enumerate(self.monitors)]

    def _browse_folder(self):
        self._suppress = True
        try:
            chosen = filedialog.askdirectory(title="저장 폴더 선택")
        finally:
            self._suppress = False
        if chosen:
            self.folder_entry.delete(0, tk.END)
            self.folder_entry.insert(0, chosen)

    def _open_folder(self):
        try:
            os.startfile(self.save_dir())
        except Exception as e:
            logging.error("Open folder failed: %s", e)

    # --- 설정 헬퍼 ----------------------------------------------------------
    def _default_monitor_index(self):
        """기본 선택 모니터: 주 모니터(보통 좌상단 (0,0))를 우선한다.

        과거엔 무조건 index 0 을 골라, 보조 모니터가 0번이면 사용자가 주 모니터에서
        한 클릭이 범위 밖이라 하나도 기록되지 않는 문제가 있었다."""
        for i, m in enumerate(self.monitors):
            if getattr(m, "is_primary", False):
                return i
        for i, m in enumerate(self.monitors):
            if m.x == 0 and m.y == 0:
                return i
        return 0

    def current_monitor(self):
        try:
            return self.monitors[self.monitor_combo.current()]
        except Exception:
            return self.monitors[0]

    def pixels_per_mm(self):
        """입력한 해상도/대각선으로 mm당 픽셀 수 계산 (mouse_distance1.0.6.py:163-173)."""
        try:
            w = int(self.res_w_entry.get())
            h = int(self.res_h_entry.get())
            diag = float(self.diag_entry.get())
            dpi = math.hypot(w, h) / diag
            return dpi / INCH_TO_MM
        except (ValueError, ZeroDivisionError):
            return DEFAULT_DPI / INCH_TO_MM

    def _autosave_interval(self):
        try:
            return max(5, int(self.autosave_entry.get()))
        except ValueError:
            return DEFAULT_AUTOSAVE_S

    def save_dir(self):
        folder = self.folder_entry.get().strip() or os.path.join(
            os.path.expanduser("~"), "Desktop", APP_DIR_NAME)
        try:
            os.makedirs(folder, exist_ok=True)
        except OSError as e:
            logging.error("makedirs failed (%s): %s", folder, e)
            folder = os.path.join(os.path.expanduser("~"), "Desktop")
        return folder

    # --- 자기 UI 클릭 제외 ---------------------------------------------------
    def _window_frame_rect(self):
        """제목표시줄·테두리까지 포함한 실제 창 사각형(GetWindowRect). 실패 시 None.

        winfo_rootx/rooty 는 제목표시줄을 뺀 클라이언트 영역이라, 그대로 쓰면 창을
        드래그하려고 제목표시줄을 누른 클릭이 제외되지 않는다."""
        try:
            hwnd = ctypes.windll.user32.GetAncestor(self.winfo_id(), 2)  # GA_ROOT
            r = wintypes.RECT()
            if ctypes.windll.user32.GetWindowRect(hwnd, ctypes.byref(r)):
                return (int(r.left), int(r.top), int(r.right), int(r.bottom))
        except Exception:
            pass
        return None

    def _update_own_rect(self):
        """우리 창의 화면 사각형(제목표시줄 포함)을 캐시. 최소화/withdraw면 None. 메인 스레드 전용."""
        try:
            if self.state() in ("iconic", "withdrawn"):
                self._own_rect = None
                return
            rect = self._window_frame_rect()
            if rect is None:   # 폴백: 클라이언트 영역 + 제목표시줄/테두리 추정 여백
                x, y = self.winfo_rootx(), self.winfo_rooty()
                w, h = self.winfo_width(), self.winfo_height()
                rect = (x - 8, y - 40, x + w + 8, y + h + 8)
            self._own_rect = rect
        except Exception:
            self._own_rect = None

    def _is_self_event(self, x, y):
        """우리 UI/다이얼로그 위의 이벤트인지(=분석에서 제외해야 하는지)."""
        if self._suppress:
            return True
        rect = self._own_rect
        return bool(rect and rect[0] <= x <= rect[2] and rect[1] <= y <= rect[3])

    # --- 녹화 제어 ----------------------------------------------------------
    def toggle_recording(self):
        if self.is_recording:
            self.stop_recording()
        else:
            self.start_recording()

    def start_recording(self):
        if self.is_recording:
            return
        # 모니터/DPI/옵션을 시작 시점에 고정(리스너 스레드는 Tk 위젯을 못 읽음)
        self._active_monitor_idx = self.monitor_combo.current()
        self._active_monitor = self.monitors[self._active_monitor_idx]
        self._ppm = self.pixels_per_mm()
        self._rec_move = bool(self.move_var.get())
        self._rec_scroll = bool(self.scroll_var.get())

        self.session_start = datetime.datetime.now()
        self.session_file = os.path.join(
            self.save_dir(), f"MouseAnalytics_{self.session_start:%Y%m%d_%H%M%S}.xlsx")

        with self.lock:
            self.click_positions.clear()
            self.click_counts = {Button.left: 0, Button.right: 0, Button.middle: 0}
            self.total_distance_mm = 0.0
            self.scroll_count = 0
            self.events.clear()
            self._last_move_pos = None
            self._last_sample_pos = None
            self._last_sample_t = 0.0
            self._win_dist = 0.0
        self._last_heatmap_png = None
        self._steps = []
        self._step_no = 1
        self._step_start = self.session_start
        self._autosave_remaining = self._autosave_interval()

        self.listener = mouse.Listener(
            on_click=self.on_click, on_move=self.on_move, on_scroll=self.on_scroll)
        try:
            self.listener.start()
        except Exception as e:
            logging.error("Mouse listener start failed: %s", e)
            messagebox.showerror("오류", f"마우스 리스너 시작 실패:\n{e}")
            self.listener = None
            return

        self.is_recording = True
        self.monitor_combo.config(state="disabled")
        self.start_button.config(text="■ 정지 (Ctrl+Shift+F9)", bg="#d9534f")
        m = self._active_monitor
        self.monitor_label.config(
            text=f"모니터  {self._active_monitor_idx} ({m.width}x{m.height})")
        logging.info("Recording started -> %s", self.session_file)
        self._grab_session_shot()      # 단계 1 배경 캡처(우리 창은 숨기고)
        self.status_label.config(text="● 녹화 중 · 단계 1", fg="#d9534f")
        self.refresh_labels()

    def stop_recording(self):
        if self.listener is not None:
            try:
                self.listener.stop()
                self.listener.join()
            except Exception as e:
                logging.error("Listener stop failed: %s", e)
            self.listener = None
        self.is_recording = False
        self.deiconify()
        self.monitor_combo.config(state="readonly")
        self.start_button.config(text="▶ 녹화 시작 (Ctrl+Shift+F9)", bg="#5cb85c")
        self.status_label.config(text="○ 대기 중", fg="gray")

        self._finalize_step()          # 마지막(현재) 단계 저장
        self.export_excel("stop")
        self.refresh_labels()
        logging.info("Recording stopped (%d step(s)).", len(self._steps))
        recorded = sum(s["left"] + s["right"] + s["middle"] + len(s["events"])
                       for s in self._steps)
        self._suppress = True
        try:
            if recorded == 0:   # 선택 모니터에서 아무 입력도 안 잡혔을 때 안내
                messagebox.showwarning(
                    "기록 없음",
                    "선택한 모니터에서 마우스 입력이 감지되지 않았습니다.\n"
                    "설정의 '모니터'가 작업 중인 화면과 같은지 확인하세요.")
            else:
                messagebox.showinfo("저장 완료", f"세션이 저장되었습니다:\n{self.session_file}")
        finally:
            self._suppress = False

    # --- 단계(소프트웨어 화면 전환) ------------------------------------------
    def _heatmap_mode(self):
        try:
            return "screenshot" if self.bg_include_var.get() else "blank"
        except Exception:
            return "blank"

    def _grab_session_shot(self):
        """현재 단계의 배경을 캡처한다. '화면 배경 포함'이 꺼져 있으면 건너뛴다.

        우리 창을 잠깐 withdraw 해 배경에 안 찍히게 하고, 실제로 사라질 시간을 둔 뒤
        선택 모니터를 캡처한다. 최소화 옵션이면 캡처 후 작업표시줄에 최소화로 둔다."""
        self._session_shot = None
        try:
            if not self.bg_include_var.get():
                return
        except Exception:
            return
        m = self._active_monitor
        try:
            self.withdraw()
            self.update()
            time.sleep(0.2)            # 창이 화면에서 실제로 사라질 시간
            self._session_shot = ImageGrab.grab(
                bbox=(m.x, m.y, m.x + m.width, m.y + m.height),
                all_screens=True).convert("RGBA")
        except Exception as e:
            logging.error("Session screenshot failed: %s", e)
        finally:
            try:
                self.iconify() if self.minimize_var.get() else self.deiconify()
            except Exception:
                self.deiconify()

    def _finalize_step(self):
        """현재 단계의 히트맵을 만들어 저장하고, 단계 통계를 기록한 뒤 카운터를 리셋한다."""
        png = None
        try:                           # 히트맵 빌드는 락 안에서 click_positions 를 스냅샷
            png = self._build_heatmap_png(self._heatmap_mode(),
                                          out_name=f"heatmap_step{self._step_no}.png")
        except Exception as e:
            logging.error("Step %d heatmap build failed: %s", self._step_no, e)
        with self.lock:
            self._steps.append({
                "no": self._step_no, "png": png,
                "left": self.click_counts[Button.left],
                "right": self.click_counts[Button.right],
                "middle": self.click_counts[Button.middle],
                "scroll": self.scroll_count,
                "distance_mm": self.total_distance_mm,
                "events": list(self.events),
                "start": self._step_start or self.session_start,
                "end": datetime.datetime.now(),
            })
            self.click_positions.clear()
            self.click_counts = {Button.left: 0, Button.right: 0, Button.middle: 0}
            self.total_distance_mm = 0.0
            self.scroll_count = 0
            self.events.clear()
            self._last_move_pos = None
            self._last_sample_pos = None
            self._last_sample_t = 0.0
            self._win_dist = 0.0
        self._step_no += 1
        self._step_start = datetime.datetime.now()

    def advance_step(self):
        """현재 단계를 저장하고 다음 단계로 넘어간다(새 배경 캡처 + 카운터 리셋)."""
        if not self.is_recording:
            return
        self._finalize_step()
        self._grab_session_shot()      # 다음 단계 화면을 새 배경으로
        self.status_label.config(text=f"● 녹화 중 · 단계 {self._step_no}", fg="#d9534f")
        self.refresh_labels()
        logging.info("Advanced to step %d", self._step_no)

    # --- 리스너 콜백 (리스너 스레드) ----------------------------------------
    def _safe(self, fn, *args):
        """콜백 예외가 리스너/프로세스를 죽이지 않도록 감싼다."""
        try:
            fn(*args)
        except Exception as e:
            logging.error("Listener callback error in %s: %s", fn.__name__, e)

    def on_click(self, x, y, button, pressed):
        self._safe(self._on_click, x, y, button, pressed)

    def _on_click(self, x, y, button, pressed):
        if not self.is_recording or not pressed:
            return
        if self._is_self_event(x, y):
            return
        m = self._active_monitor
        if not (m.x <= x < m.x + m.width and m.y <= y < m.y + m.height):
            return
        rx, ry = x - m.x, y - m.y
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
        with self.lock:
            if button in self.click_counts:
                self.click_counts[button] += 1
            self.click_positions.append((rx, ry))
            if len(self.click_positions) > POINT_CAP:
                self.click_positions.pop(0)
            self.events.append((ts, "click", button.name, rx, ry,
                                self._active_monitor_idx, ""))
            if len(self.events) > POINT_CAP:
                self.events.pop(0)
        self.after(0, self.refresh_labels)

    def on_move(self, x, y):
        self._safe(self._on_move, x, y)

    def _on_move(self, x, y):
        if not self.is_recording or not self._rec_move:
            return
        if self._is_self_event(x, y):
            return
        now = time.monotonic()
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
        m = self._active_monitor
        with self.lock:
            if self._last_move_pos is not None:
                d = math.hypot(x - self._last_move_pos[0], y - self._last_move_pos[1])
                self.total_distance_mm += d / self._ppm   # 거리는 매 이벤트 누적
                self._win_dist += d
            self._last_move_pos = (x, y)

            if self._last_sample_pos is None:
                moved = float("inf")
            else:
                moved = math.hypot(x - self._last_sample_pos[0],
                                   y - self._last_sample_pos[1])
            # 코얼레싱: 100ms 경과 또는 50px 이동 시에만 샘플 1행 기록
            if (now - self._last_sample_t) >= MOVE_MIN_INTERVAL_S or moved >= MOVE_MIN_DIST_PX:
                if m.x <= x < m.x + m.width and m.y <= y < m.y + m.height:
                    self.events.append((ts, "move", "", x - m.x, y - m.y,
                                        self._active_monitor_idx, round(self._win_dist, 1)))
                    if len(self.events) > POINT_CAP:
                        self.events.pop(0)
                self._last_sample_pos = (x, y)
                self._last_sample_t = now
                self._win_dist = 0.0
        self.after(0, self.refresh_labels)

    def on_scroll(self, x, y, dx, dy):
        self._safe(self._on_scroll, x, y, dx, dy)

    def _on_scroll(self, x, y, dx, dy):
        if not self.is_recording or not self._rec_scroll:
            return
        if self._is_self_event(x, y):
            return
        m = self._active_monitor
        if not (m.x <= x < m.x + m.width and m.y <= y < m.y + m.height):
            return
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
        with self.lock:
            self.scroll_count += abs(int(dy))
            self.events.append((ts, "scroll", "", x - m.x, y - m.y,
                                self._active_monitor_idx, ""))
            if len(self.events) > POINT_CAP:
                self.events.pop(0)
        self.after(0, self.refresh_labels)

    # --- 라이브 UI 갱신 (메인 스레드) ---------------------------------------
    def refresh_labels(self):
        with self.lock:
            left = self.click_counts[Button.left]
            right = self.click_counts[Button.right]
            middle = self.click_counts[Button.middle]
            dist_mm = self.total_distance_mm
            scrolls = self.scroll_count
        total = left + right + middle
        px = dist_mm * self._ppm
        cm = dist_mm * MM_TO_CM
        self.clicks_label.config(
            text=f"클릭  총 {total}  (좌 {left} · 우 {right} · 휠 {middle})")
        self.distance_label.config(text=f"이동 거리  {px:,.0f} px  /  {cm:.1f} cm")
        self.scroll_label.config(text=f"스크롤  {scrolls} 칸")

    def tick(self):
        """1초마다: 경과시간 갱신 + 자동저장 카운트다운 + 자기영역 안전망 갱신."""
        self._update_own_rect()
        if self.is_recording and self.session_start is not None:
            elapsed = (datetime.datetime.now() - self.session_start).total_seconds()
            self.time_label.config(text=f"경과 시간  {self._fmt_hms(elapsed)}")
            self._autosave_remaining -= 1
            if self._autosave_remaining <= 0:
                self.autosave()
                self._autosave_remaining = self._autosave_interval()
        self.after(1000, self.tick)

    def autosave(self):
        if self.is_recording and self.autosave_var.get():
            self.export_excel("autosave")

    @staticmethod
    def _fmt_hms(seconds):
        s = int(seconds)
        return f"{s // 3600:02d}:{(s % 3600) // 60:02d}:{s % 60:02d}"

    # --- 히트맵 -------------------------------------------------------------
    def _build_heatmap_png(self, mode, out_name="heatmap.png"):
        """클릭 위치 히트맵 PNG 생성 (mouse_click_move2.py:275-324 재사용).

        mode='screenshot'면 화면 캡처 위에, 'blank'면 흰 캔버스 위에 합성한다.
        out_name 으로 단계별 파일명을 분리한다.
        """
        monitor = self._active_monitor or self.current_monitor()
        if mode == "screenshot" and self._session_shot is not None:
            # 녹화 시작 때 캡처해 둔 세션 배경 사용(우리 창이 안 찍히고 안정적)
            background = self._session_shot.copy()
            actual_w, actual_h = background.size
        elif mode == "screenshot":
            # 세션 캡처가 없으면(녹화 중이 아닐 때 등) 즉시 캡처. ImageGrab bbox 는
            # (left, top, right, bottom), all_screens=True 로 보조/음수 좌표까지 정확히.
            bbox = (monitor.x, monitor.y,
                    monitor.x + monitor.width, monitor.y + monitor.height)
            self._suppress = True   # 캡처 순간의 자기 이벤트 무시
            try:
                background = ImageGrab.grab(bbox=bbox, all_screens=True).convert("RGBA")
            finally:
                self._suppress = False
            actual_w, actual_h = background.size
        else:
            actual_w, actual_h = monitor.width, monitor.height
            background = Image.new("RGBA", (actual_w, actual_h), (255, 255, 255, 255))

        if actual_w <= 0 or actual_h <= 0:   # 방어: 잘못된 화면 크기
            raise ValueError(f"화면 크기가 잘못됨({actual_w}x{actual_h}) — 모니터 선택을 확인하세요.")

        with self.lock:
            points = list(self.click_positions)

        # 4K 빈 캔버스는 float32 ~33MB. 버튼/정지 시에만 호출되므로 허용.
        heatmap_data = np.zeros((actual_h, actual_w), dtype=np.float32)
        for px, py in points:
            ix, iy = int(px), int(py)
            if 0 <= ix < actual_w and 0 <= iy < actual_h:
                heatmap_data[iy, ix] += 1

        # 가우시안 블러를 float 로 수행(희소 클릭도 보존). 그 뒤 정규화.
        blurred = _gaussian_blur(heatmap_data, GAUSS_SIGMA)
        peak = float(blurred.max())
        if peak > 0:                       # 빈 세션 NaN 방지
            blurred = blurred / peak

        # 색: turbo 컬러맵(부드러운 그라데이션) + 밀도 비례 알파(빈 곳 투명, 핫스팟 진하게)
        norm = np.clip(blurred, 0.0, 1.0)
        rgb = _turbo_rgb(norm)
        colored = np.empty((actual_h, actual_w, 4), dtype=np.uint8)
        colored[..., :3] = (rgb * 255).astype(np.uint8)
        colored[..., 3] = ((norm ** HEAT_ALPHA_GAMMA) * HEAT_MAX_ALPHA * 255).astype(np.uint8)
        heatmap_image = Image.fromarray(colored, mode="RGBA")

        # 옅은 파란 베이스 한 겹(슬라이더로 조절) → 배경 위·히트맵 아래에 합성
        try:
            base_a = float(self.blue_base_var.get())
        except Exception:
            base_a = BLUE_BASE_DEFAULT
        if base_a > 0:
            base = Image.new("RGBA", (actual_w, actual_h),
                             BLUE_BASE_RGB + (int(np.clip(base_a, 0.0, 1.0) * 255),))
            background = Image.alpha_composite(background, base)

        combined = Image.alpha_composite(background, heatmap_image)
        png_path = os.path.join(self.temp_dir, out_name)
        combined.convert("RGB").save(png_path, format="PNG")
        return png_path

    # --- 엑셀 저장 ----------------------------------------------------------
    def export_excel(self, reason):
        """summary(단계별 표) + 단계별 히트맵 시트(step1, step2…) + events 를 저장."""
        if not self.session_start:
            self.session_start = datetime.datetime.now()
        if not self.session_file:
            self.session_file = os.path.join(
                self.save_dir(),
                f"MouseAnalytics_{self.session_start:%Y%m%d_%H%M%S}.xlsx")

        steps = list(self._steps)                       # 완료된 단계
        with self.lock:                                  # 진행 중인 현재 단계 스냅샷
            cur = {
                "left": self.click_counts[Button.left],
                "right": self.click_counts[Button.right],
                "middle": self.click_counts[Button.middle],
                "scroll": self.scroll_count,
                "distance_mm": self.total_distance_mm,
                "n_events": len(self.events),
            }
        end = datetime.datetime.now()
        monitor = self._active_monitor or self.current_monitor()

        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            self._write_summary_sheet(wb, steps, cur, self.session_start, end, monitor)
            for rec in steps:
                ws = wb.create_sheet(f"step{rec['no']}")
                if rec["png"] and os.path.exists(rec["png"]):
                    try:
                        ws.add_image(XLImage(rec["png"]), "A1")
                    except Exception as e:
                        logging.error("Embed step%d image failed: %s", rec["no"], e)
            self._write_events_sheet(wb, steps)
            self._atomic_save(wb, reason)
        except Exception as e:
            logging.error("Excel export failed (%s): %s", reason, e)
            if reason != "autosave":
                self._suppress = True
                try:
                    messagebox.showerror("오류", f"엑셀 저장 실패:\n{e}")
                finally:
                    self._suppress = False

    def _write_summary_sheet(self, wb, steps, cur, start, end, monitor):
        ws = wb.create_sheet("summary")
        dpi = self._ppm * INCH_TO_MM
        try:
            mon_idx = self.monitors.index(monitor)
        except ValueError:
            mon_idx = self._active_monitor_idx
        bg = "화면 캡처" if self._heatmap_mode() == "screenshot" else "빈 캔버스"
        cur_active = (cur["left"] + cur["right"] + cur["middle"] + cur["n_events"]) > 0

        info = [
            ("세션 시작", start.strftime("%Y-%m-%d %H:%M:%S")),
            ("세션 종료", end.strftime("%Y-%m-%d %H:%M:%S")),
            ("지속 시간", self._fmt_hms((end - start).total_seconds())),
            ("모니터", f"{mon_idx}: {monitor.width}x{monitor.height} at ({monitor.x},{monitor.y})"),
            ("사용 DPI", round(dpi, 1)),
            ("히트맵 배경", bg),
            ("단계 수", len(steps) + (1 if cur_active else 0)),
        ]
        r = 1
        for label, value in info:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=value)
            r += 1
        r += 1                                          # 빈 줄

        header = ["단계", "좌클릭", "우클릭", "휠클릭", "총클릭", "스크롤(칸)", "이동(cm)", "지속"]
        for c, h in enumerate(header, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = openpyxl.styles.Font(bold=True)
        r += 1

        tot = {"left": 0, "right": 0, "middle": 0, "scroll": 0, "dist": 0.0}

        def step_row(rownum, label, left, right, middle, scroll, dist_mm, dur):
            vals = [label, left, right, middle, left + right + middle,
                    scroll, round(dist_mm * MM_TO_CM, 1), dur]
            for c, v in enumerate(vals, start=1):
                ws.cell(row=rownum, column=c, value=v)

        for rec in steps:
            step_row(r, rec["no"], rec["left"], rec["right"], rec["middle"],
                     rec["scroll"], rec["distance_mm"],
                     self._fmt_hms((rec["end"] - rec["start"]).total_seconds()))
            tot["left"] += rec["left"]; tot["right"] += rec["right"]
            tot["middle"] += rec["middle"]; tot["scroll"] += rec["scroll"]
            tot["dist"] += rec["distance_mm"]
            r += 1
        if cur_active:
            step_row(r, "현재(진행중)", cur["left"], cur["right"], cur["middle"],
                     cur["scroll"], cur["distance_mm"], "")
            tot["left"] += cur["left"]; tot["right"] += cur["right"]
            tot["middle"] += cur["middle"]; tot["scroll"] += cur["scroll"]
            tot["dist"] += cur["distance_mm"]
            r += 1

        step_row(r, "합계", tot["left"], tot["right"], tot["middle"],
                 tot["scroll"], tot["dist"], "")
        for c in range(1, 9):
            ws.cell(row=r, column=c).font = openpyxl.styles.Font(bold=True)

        ws.column_dimensions["A"].width = 14
        for col in "BCDEFGH":
            ws.column_dimensions[col].width = 10

    def _write_events_sheet(self, wb, steps):
        ws = wb.create_sheet("events")
        ws.append(["step", "timestamp", "event_type", "button", "x", "y",
                   "monitor_index", "distance_delta_px"])
        for rec in steps:
            for row in rec["events"]:
                ws.append([rec["no"]] + list(row))

    def _atomic_save(self, wb, reason):
        """temp 파일에 쓴 뒤 원자적 교체. 원본이 잠겨 있으면 백업본으로 저장."""
        final = self.session_file
        tmp = final + ".tmp"
        wb.save(tmp)
        try:
            os.replace(tmp, final)
            logging.info("Excel saved (%s): %s", reason, final)
        except (PermissionError, OSError):
            stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            backup = os.path.join(os.path.dirname(final),
                                  f"MouseAnalytics_backup_{stamp}.xlsx")
            os.replace(tmp, backup)
            logging.warning("Final locked; saved backup: %s", backup)
            if reason != "autosave":
                self._suppress = True
                try:
                    messagebox.showwarning(
                        "파일 잠금",
                        f"원본 파일이 열려 있어 백업으로 저장했습니다:\n{backup}")
                finally:
                    self._suppress = False

    # --- 종료 ---------------------------------------------------------------
    def on_closing(self):
        if self.listener is not None:
            try:
                self.listener.stop()
                self.listener.join()
            except Exception as e:
                logging.error("Listener stop on close failed: %s", e)
            self.listener = None
        # 녹화 중이었다면 현재 단계를 마무리하고 최종 저장
        if self.is_recording:
            self.is_recording = False
            try:
                self._finalize_step()
                self.export_excel("close")
            except Exception as e:
                logging.error("Final export on close failed: %s", e)
        if self.keyboard_listener is not None:
            try:
                self.keyboard_listener.stop()
                self.keyboard_listener.join()
            except Exception as e:
                logging.error("Keyboard listener stop failed: %s", e)
        self.destroy()
        try:
            if os.path.isdir(self.temp_dir):
                shutil.rmtree(self.temp_dir)
        except Exception as e:
            logging.warning("Temp dir cleanup failed (%s): %s", self.temp_dir, e)


def main():
    root = MouseAnalytics()
    root.protocol("WM_DELETE_WINDOW", root.on_closing)
    root.mainloop()


if __name__ == "__main__":
    main()
